#==================================================================================
#===    Author:            ANDRES OBREGON                                       ===
#===    Create date:       25/04/2023                                           ===
#===    Description:       Reporte Facturación més/Comerciales                  ===
#==================================================================================

import tkinter as tk
import tkinter.font as tkEstile
import datetime as dt
import datetime 
import pyodbc
import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Alignment
from tkinter import filedialog
from tkinter import ttk
from tkinter import *
from datetime import datetime, date
from PIL import ImageTk, Image

#=====================VARIABLES GLOBALES=================================================

fechmin = "none"            #FECHA MAS BAJA
fechmax = "none"            #FECHA MAS ALTA
listNombreComer = ["TODOS"]        #LISTA COMERCIALES
archivo = ""                #UBICAION DE ARCHIVO A GUARDAR
comercial = ""              #COMERCIAL SELECIONADO

#=====================DATOS SERVIDOR=====================================================

server = ""
database = ""
username = ""
password = ""

#=====================Consultar Lista de comerciales=====================================

def conectarBDComerciales():                                #Con esta funcion hacemos consulta al servidor la lista de comerciales y la reflemos en la variable para que se ven en desplegable
    global listNombreComer
    consultaComrcial = "select Nombre from Comerciales"
    if len(listNombreComer) == 1 : 
        cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cursor = cnxn.cursor()
        cursor.execute(consultaComrcial)
        for row in cursor.fetchall():
            listNombreComer.append(row[0])
        cursor.close()
        cnxn.close()
        return listNombreComer

#=========================Cosultar Lineas cosulta ==========================================

def cosultarLinea():                                       #Aqui consultamos sin entre las fechas selecionadas hay datos
    global fechmin
    global fechmax
    consultaComrcialR = ""
    if cb.get() == 1:
        consultaComrcialR = "".format(fechmin,fechmax,comercial)
    else:
        consultaComrcialR = """
        """.format(fechmin,fechmax)
    cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    cursor.execute(consultaComrcialR)
    numero =  cursor.fetchone()[0]
    if comercial == "TODOS":
        numero = 1
    cursor.close()
    cnxn.close()
    return numero

#==============================Consulta Archivo=============================================

def seleccionar_archivo():                                 #Preguntamos donde quieres guardar el EXEL
    global archivo                                         #Varible de archivo
    fecha_actual = datetime.now()
    fecha_formateada = fecha_actual.strftime('%d-%m-%y')
    tipos_archivo = [("Archivos de Excel","*.xlsx")]
    nombre_archivo = "Facturacion"+comercial+fecha_formateada+".xlsx"
    # Abre el cuadro de diálogo de guardar archivo con el nombre de archivo inicial
    archivo = filedialog.asksaveasfilename(initialfile=nombre_archivo, filetypes=tipos_archivo)
    return archivo

#===================================Generar Exel=============================================

def generarExel():
    consulta = ""
    if cb.get() == 1:    
        if comercial == "TODOS":
            consulta = """
            """.format(fechmin,fechmax,comercial)
        else :
            consulta = """
            """.format(fechmin,fechmax,comercial)
    else:
         consulta = """
        """.format(fechmin,fechmax)
    cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    cursor.execute(consulta)
    row = cursor.fetchall()

    libro = openpyxl.Workbook()
    hoja_excel = libro.active
    hoja_excel.title = 'Facturación'
    # Crea encabezados para las columnas
    hoja_excel.cell(row=1, column=1, value="Fecha").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=2, value="Numero").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=3, value="Nombre").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=4, value="NIF").alignment = Alignment(horizontal="center")
    if cb.get() == 1:
        hoja_excel.cell(row=1, column=5, value="Descripción").alignment = Alignment(horizontal="center")    #DESCRIPCION
        hoja_excel.cell(row=1, column=6, value="Importe").alignment = Alignment(horizontal="center")        #IMPORTE
        hoja_excel.cell(row=1, column=7, value="Comercial").alignment = Alignment(horizontal="center")      #COMERCIAL
    else: 
        hoja_excel.cell(row=1, column=5, value="Importe").alignment = Alignment(horizontal="center")        #IMPORTE
        hoja_excel.cell(row=1, column=6, value="VTO").alignment = Alignment(horizontal="center")            #VTO
    fila = 2

    for x in row: # Escribimos los datos del query en el exel
        hoja_excel.cell(row=fila, column=1, value=x[0])
        hoja_excel.cell(row=fila, column=2, value=x[1])
        hoja_excel.cell(row=fila, column=3, value=x[2])
        hoja_excel.cell(row=fila, column=4, value=x[3])
        if cb.get() == 1:
            hoja_excel.cell(row=fila, column=5, value=x[4])             #DESCRIPCION
            hoja_excel.cell(row=fila, column=6, value=x[5])             #IMPORTE
            hoja_excel.cell(row=fila, column=7, value=x[6])             #COMERCIAL
        else: 
            hoja_excel.cell(row=fila, column=5, value=x[4])             #IMPORTE
            hoja_excel.cell(row=fila, column=6, value=x[5])             #VTO
        fila += 1
        
    ajustarExel(hoja_excel)
    libro.save(archivo)                             # Guarda el archivo de Excel
    cursor.close()
    cnxn.close()

#==============================Ajustar Exels===============================================

def ajustarExel(hoja_excel):                            #Con esto le damos formato a las celdas para que quedebe bien
    for col in hoja_excel.columns: 
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja_excel.column_dimensions[column].width = adjusted_width

#==============================Validar fechas===============================================

def comprbarFecha():
    global fechmin
    global fechmax
    fecha_ingresadamin = fechaminima.get()
    fecha_ingresadamax = fechamaxima.get()
    try:
        # Intenta convertir la fecha ingresada en un objeto datetime
        fechaM1 = datetime.strptime(fecha_ingresadamin, "%d/%m/%Y")        
        fechaX1 = datetime.strptime(fecha_ingresadamax, "%d/%m/%Y")
        # Comprueba si la fecha es menor o igual a la fecha actual
        if fechaM1 <= fechaX1:
            fechmin = datetime.strftime(fechaM1, "%d/%m/%Y")        
            fechmax = datetime.strftime(fechaX1, "%d/%m/%Y")
            if cb.get() == 1:                                                         #ON SELECECION POR COMERCIAL
                resultado.config(text="") 
                comeSelecion()
            elif cb.get() == 0:                                                       #OFF SELECECION POR COMERCIAL
                if cosultarLinea() == 0:                                              #Comprobar sin la consulta esta vacia y generar error
                    resultado.config(text="Periodo sin facturación")
                    resultado.place(x=180, y=250)
                else :
                    seleccionar_archivo()
                    generarExel()

            return fechmin,fechmax
        else:
            resultado.config(text="La fecha maxima no puede ser menor que la fecha minima")
            resultado.place(x=35, y=250)

    except ValueError:
        # Si la conversión falla, significa que la fecha no está en el formato correcto
        resultado.config(text="La fecha no está en el formato correcto")
        resultado.place(x=115, y=250)

#======================Validar Comercial Introducido============================================

def comeSelecion():
    global comercial
    comercial = combo.get()
    if comercial == "":
        resultado.config(text="Selecione un comercial")
        resultado.place(x=160, y=250)
    else:
        #Aqui deberiamos ejecutar desde DEF
        if cosultarLinea() == 0:                                              #Comprobar sin la consulta esta vacia y generar error
            resultado.config(text="Comercial sin descuentos en este periodo")
            resultado.place(x=100, y=250)
        else:
            seleccionar_archivo()
            generarExel()
            
#==============================Activar y Buscar comercial=========================================

def isChecked():
    global comercial
    conectarBDComerciales()
    if cb.get() == 1: 
        combo['state'] = NORMAL
        combo['values'] = listNombreComer
    elif cb.get() == 0: 
        comercial = ""
        combo['state'] = DISABLED
    else: print('')

#=========================================GRAFICO======================================================

ventana = tk.Tk()
ventana.iconbitmap('')
fontStyleVersion = tkEstile.Font(family="arial", size=7)
fontStyle = tkEstile.Font(family="Impact", size=20)

                                                                    # Cargar la imagen
image = Image.open('')
resized_image = image.resize((100, 35))
photo = ImageTk.PhotoImage(resized_image)
label = tk.Label(ventana, image=photo)
label.place(x=5, y=5)
                                                                    #Version de programa 
versionP = tk.Label(master=ventana,text='V.1 25/04/2023', font=fontStyleVersion)
versionP.place(x=420, y=5)
                                                                    #Titulo Facturacion
titulo = tk.Label(master=ventana,text='Generador facturación', font=fontStyle, pady= 15)
titulo.place(x=124, y=5)
                                                                    #Texto Inicio de periodo
fechamin = tk.Label(master=ventana, text="Inicio periodo:  ")
fechamin.place(x=125, y=75)
                                                                    #Campo de entrada Fecha minima
fechaminima = tk.Entry(master=ventana)
fechaminima.place(x=250, y=75)
                                                                    #Texto Fin de periodo
fechamax = tk.Label(master=ventana, text="Fin periodo:  ")
fechamax.place(x=125, y=100)
                                                                    #Campo de entrada Fecha Maxima
fechamaxima = tk.Entry(master=ventana)
fechamaxima.place(x=250, y=100)
                                                                    #Boton de Validar fechas y comercial
boton_validar = tk.Button(master=ventana, text="Validar", command=comprbarFecha)
boton_validar.place(x=229, y=150)

cb = IntVar()                                                       #Check para filtrar por Comercial
check = tk.Checkbutton(text="Descuentos por comercial", variable=cb, onvalue=1, offvalue=0, command=isChecked)
check.place(x=250, y=200)

combo = ttk.Combobox(values=listNombreComer)                        #Lista de comerciales 
combo['state'] = DISABLED
combo.place(x=100, y=200)

resultado = tk.Label(ventana, text="",fg='#f00',font=10)            #Mensaje de error 

ventana.title("Generador Facturación")
ventana.minsize(500,300)
ventana.maxsize(550,350)
ventana.mainloop()

 #ventana.destroy()